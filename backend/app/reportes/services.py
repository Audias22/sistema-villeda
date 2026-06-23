import os
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app import db
from app.expedientes.models import Expediente
from app.documentos.models import Documento
from app.clientes.models import Cliente
from app.busquedas.models import Busqueda
from app.common.models import AreaJuridica, EstadoExpediente, TipoExpediente, Prioridad
from app.usuarios.models import Usuario
from sqlalchemy import func


RUTA_EXPORTACIONES = os.path.join(os.path.dirname(__file__), '..', '..', 'almacenamiento', 'exportaciones')


def obtener_dashboard():
    """
    Estadísticas generales del sistema para el dashboard principal.
    Combina datos de expedientes, documentos, clientes y búsquedas (TBR).
    """

    total_expedientes = Expediente.query.count()
    total_documentos = Documento.query.count()
    total_clientes = Cliente.query.filter_by(activo=True).count()
    total_busquedas = Busqueda.query.count()

    expedientes_por_area = db.session.query(
        AreaJuridica.nombre,
        func.count(Expediente.id_expediente).label('total')
    ).outerjoin(
        Expediente, Expediente.id_area == AreaJuridica.id_area
    ).group_by(AreaJuridica.id_area, AreaJuridica.nombre).all()

    expedientes_por_estado = db.session.query(
        EstadoExpediente.nombre,
        func.count(Expediente.id_expediente).label('total')
    ).outerjoin(
        Expediente, Expediente.id_estado == EstadoExpediente.id_estado
    ).group_by(EstadoExpediente.id_estado, EstadoExpediente.nombre).all()

    metricas_tbr = db.session.query(
        func.avg(Busqueda.tiempo_respuesta_ms).label('promedio_ms'),
        func.min(Busqueda.tiempo_respuesta_ms).label('minimo_ms'),
        func.max(Busqueda.tiempo_respuesta_ms).label('maximo_ms')
    ).first()

    documentos_duplicados = Documento.query.filter_by(es_duplicado_exacto=True).count()

    return {
        'totales': {
            'expedientes': total_expedientes,
            'documentos':  total_documentos,
            'clientes':    total_clientes,
            'busquedas':   total_busquedas
        },
        'expedientes_por_area': [
            {'area': nombre, 'total': total} for nombre, total in expedientes_por_area
        ],
        'expedientes_por_estado': [
            {'estado': nombre, 'total': total} for nombre, total in expedientes_por_estado
        ],
        'tbr': {
            'promedio_ms': round(metricas_tbr.promedio_ms, 2) if metricas_tbr.promedio_ms else 0,
            'minimo_ms':   metricas_tbr.minimo_ms or 0,
            'maximo_ms':   metricas_tbr.maximo_ms or 0
        },
        'documentos_duplicados': documentos_duplicados
    }


def exportar_expedientes_excel(id_area=None, id_estado=None):
    """
    Genera un archivo Excel con el listado completo de expedientes,
    combinando datos de cliente, área, tipo, estado, prioridad,
    usuario asignado y conteo de documentos.
    """

    query = db.session.query(
        Expediente.numero_expediente,
        Expediente.titulo,
        Cliente.primer_nombre,
        Cliente.primer_apellido,
        Cliente.razon_social,
        AreaJuridica.nombre.label('area'),
        TipoExpediente.nombre.label('tipo'),
        EstadoExpediente.nombre.label('estado'),
        Prioridad.nombre.label('prioridad'),
        Expediente.fecha_apertura,
        Usuario.nombre.label('asignado_nombre'),
        Usuario.apellido.label('asignado_apellido')
    ).join(
        Cliente, Expediente.id_cliente == Cliente.id_cliente
    ).join(
        AreaJuridica, Expediente.id_area == AreaJuridica.id_area
    ).join(
        TipoExpediente, Expediente.id_tipo_expediente == TipoExpediente.id_tipo
    ).join(
        EstadoExpediente, Expediente.id_estado == EstadoExpediente.id_estado
    ).join(
        Prioridad, Expediente.prioridad == Prioridad.id_prioridad
    ).join(
        Usuario, Expediente.id_usuario_asignado == Usuario.id_usuario
    )

    if id_area:
        query = query.filter(Expediente.id_area == id_area)

    if id_estado:
        query = query.filter(Expediente.id_estado == id_estado)

    filas = query.order_by(Expediente.fecha_apertura.desc()).all()

    conteo_documentos = dict(
        db.session.query(
            Documento.id_expediente,
            func.count(Documento.id_documento)
        ).group_by(Documento.id_expediente).all()
    )

    expedientes_ids = {
        e.numero_expediente: e.id_expediente
        for e in Expediente.query.with_entities(Expediente.numero_expediente, Expediente.id_expediente).all()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Expedientes"

    encabezados = [
        'No. Expediente', 'Título', 'Cliente', 'Área', 'Tipo',
        'Estado', 'Prioridad', 'Fecha Apertura', 'Documentos', 'Asignado a'
    ]

    color_encabezado = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fuente_encabezado = Font(color="FFFFFF", bold=True, size=11)

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        celda.fill = color_encabezado
        celda.font = fuente_encabezado
        celda.alignment = Alignment(horizontal='center', vertical='center')

    for fila_idx, fila in enumerate(filas, start=2):
        nombre_cliente = fila.razon_social if fila.razon_social else f"{fila.primer_nombre or ''} {fila.primer_apellido or ''}".strip()
        id_exp = expedientes_ids.get(fila.numero_expediente)
        total_docs = conteo_documentos.get(id_exp, 0)
        asignado = f"{fila.asignado_nombre} {fila.asignado_apellido}"

        ws.cell(row=fila_idx, column=1, value=fila.numero_expediente)
        ws.cell(row=fila_idx, column=2, value=fila.titulo)
        ws.cell(row=fila_idx, column=3, value=nombre_cliente)
        ws.cell(row=fila_idx, column=4, value=fila.area)
        ws.cell(row=fila_idx, column=5, value=fila.tipo)
        ws.cell(row=fila_idx, column=6, value=fila.estado)
        ws.cell(row=fila_idx, column=7, value=fila.prioridad)
        ws.cell(row=fila_idx, column=8, value=fila.fecha_apertura.strftime('%d/%m/%Y') if fila.fecha_apertura else '')
        ws.cell(row=fila_idx, column=9, value=total_docs)
        ws.cell(row=fila_idx, column=10, value=asignado)

    anchos = [16, 35, 25, 12, 22, 14, 10, 15, 12, 20]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}{len(filas) + 1}"
    ws.freeze_panes = "A2"

    os.makedirs(RUTA_EXPORTACIONES, exist_ok=True)
    nombre_archivo = f"expedientes_{uuid.uuid4().hex[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    ruta_completa = os.path.join(RUTA_EXPORTACIONES, nombre_archivo)

    wb.save(ruta_completa)

    return ruta_completa, nombre_archivo